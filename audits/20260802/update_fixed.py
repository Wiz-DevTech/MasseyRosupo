#!/usr/bin/env python3
"""Update the Fixed column in the master audit CSV + regenerate XLSX (2026-08-02 remediation)."""
import csv, os, sys

OUT = "/opt/masseyrosupo.com/audits/20260802/reports"
CSV = os.path.join(OUT, "master-audit-20260802.csv")

DONE = {
    "P0-01": "Yes", "P0-04": "Yes", "P0-05": "Yes",
    "P1-06": "Yes", "P1-07": "Yes", "P1-09": "Yes", "P1-10": "Yes",
    "P1-11": "Yes", "P1-14": "Yes", "P1-15": "Yes",
    "P2-16": "Yes", "P2-17": "Yes", "P2-19": "Yes",
    "P3-24": "Yes", "P3-26": "Yes",
    "P0-03": "Yes",  # origin + code + nginx + DNS record created + public TLS verified 2026-08-02
    # Deferred-items pass (2026-08-02):
    "P1-12": "Yes", "P1-13": "Yes", "P2-18": "Yes", "P2-20": "Yes", "P2-21": "Yes",
    "P3-22": "Yes", "P3-23": "Yes", "P3-25": "Yes",
    "DI-2": "Yes", "WF1-1": "Yes", "BIZ-5": "Yes", "BIZ-6": "Yes",
    "NOTIF-1": "Yes", "NOTIF-2": "Yes", "SEC-2": "Yes", "SEC-3": "Yes", "SYNC-1": "Yes",
}

rows = list(csv.reader(open(CSV, encoding="utf-8-sig")))
hdr = rows[0]
fixed_idx = hdr.index("Fixed")
id_idx = hdr.index("Finding ID")
updated = 0
for r in rows[1:]:
    fid = r[id_idx].strip()
    if fid in DONE:
        r[fixed_idx] = DONE[fid]
        updated += 1

with open(CSV, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerows(rows)
print(f"CSV updated: {updated} rows marked in Fixed column")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Audit"
    ws.append(hdr)
    sev_fill = {
        "Critical": PatternFill("solid", fgColor="FFC7CE"),
        "High": PatternFill("solid", fgColor="FFEB9C"),
        "Medium": PatternFill("solid", fgColor="FFEFD5"),
        "Low": PatternFill("solid", fgColor="E2EFDA"),
        "Info": PatternFill("solid", fgColor="DDEBF7"),
    }
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
    for r in rows[1:]:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        if row[6].value in sev_fill:
            row[6].fill = sev_fill[row[6].value]
    widths = [16,12,38,40,70,10,10,16,9,40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    xlsx = os.path.join(OUT, "master-audit-20260802.xlsx")
    wb.save(xlsx)
    print(f"XLSX regenerated: {xlsx}")
except ImportError:
    print("openpyxl unavailable - CSV only")

# Summary of Fixed states
from collections import Counter
states = Counter(r[fixed_idx] for r in rows[1:])
print("Fixed column distribution:", dict(states))
